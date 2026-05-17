# =============================================================================
# FLUXAGENT V2 — PIPELINE DE INGESTA (RAG)
# =============================================================================
# Orquesta el procesamiento de documentos para construir la base de conocimiento
# vectorial de cada agente. Flujo:
#   Archivo/URL → Extracción de texto → Chunking → Embedding → PostgreSQL
#
# Fuentes soportadas:
#   - PDF     : catálogos, manuales, precios
#   - Excel   : inventarios, listas de productos
#   - URL     : páginas web de la empresa
# =============================================================================

import asyncio
import hashlib
import json
import logging
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from textwrap import wrap
from typing import Optional
from uuid import UUID

import httpx
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from config import obtener_config

logger = logging.getLogger(__name__)
config = obtener_config()

# Tamaño máximo de cada fragmento en caracteres (~200 tokens aprox.)
CHUNK_SIZE = 800
# Solapamiento entre fragmentos para no perder contexto en los bordes
CHUNK_OVERLAP = 100


# =============================================================================
# ESTRUCTURAS DE DATOS
# =============================================================================

@dataclass
class Chunk:
    """Representa un fragmento de texto listo para vectorizar."""
    contenido:    str
    fuente_nombre: str
    fuente_tipo:  str
    fuente_url:   Optional[str] = None
    pagina_numero: Optional[int] = None
    orden_chunk:  int = 0


# =============================================================================
# EXTRACCIÓN DE TEXTO
# =============================================================================

def extraer_texto_pdf(contenido: bytes, nombre_archivo: str) -> list[Chunk]:
    """
    Extrae texto de un PDF página por página y lo divide en chunks.

    Estrategia:
      - Procesa cada página de forma independiente
      - Respeta los saltos de página como límites naturales de contexto
      - Aplica chunking con solapamiento para no perder contexto

    Args:
        contenido        : Bytes del archivo PDF
        nombre_archivo   : Nombre original del archivo (para metadatos)

    Returns:
        Lista de Chunk listos para vectorizar
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        raise RuntimeError("pypdf no instalado. Añadir 'pypdf' a requirements.txt")

    chunks: list[Chunk] = []
    reader = PdfReader(BytesIO(contenido))

    logger.info(f"Procesando PDF '{nombre_archivo}': {len(reader.pages)} páginas")

    for num_pagina, pagina in enumerate(reader.pages, start=1):
        texto = pagina.extract_text() or ""
        texto = _limpiar_texto(texto)

        if not texto.strip():
            continue

        fragmentos = _dividir_en_chunks(texto)
        for sort_order, fragmento in enumerate(fragmentos):
            chunks.append(Chunk(
                contenido=fragmento,
                fuente_nombre=nombre_archivo,
                fuente_tipo="pdf",
                pagina_numero=num_pagina,
                orden_chunk=sort_order,
            ))

    logger.info(f"PDF '{nombre_archivo}': {len(chunks)} chunks generados")
    return chunks


def extraer_texto_excel(contenido: bytes, nombre_archivo: str) -> list[Chunk]:
    """
    Extrae texto de un archivo Excel convirtiendo cada hoja en texto narrativo.

    Estrategia:
      - Cada fila de datos se convierte en una oración descriptiva
      - Los encabezados se usan como etiquetas de contexto
      - Ideal para catálogos de productos con columnas Nombre/Precio/Descripción

    Args:
        contenido       : Bytes del archivo Excel (.xlsx)
        nombre_archivo  : Nombre original del archivo

    Returns:
        Lista de Chunk listos para vectorizar
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no instalado.")

    chunks: list[Chunk] = []
    wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)

    logger.info(f"Procesando Excel '{nombre_archivo}': {len(wb.sheetnames)} hojas")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        filas = list(hoja.iter_rows(values_only=True))

        if not filas:
            continue

        # La primera fila son los encabezados
        encabezados = [str(c).strip() if c else f"columna_{i}" for i, c in enumerate(filas[0])]
        texto_acumulado = []

        for num_fila, fila in enumerate(filas[1:], start=2):
            # Construir texto narrativo: "Nombre: X, Precio: Y, Stock: Z"
            partes = []
            for encabezado, valor in zip(encabezados, fila):
                if valor is not None and str(valor).strip():
                    partes.append(f"{encabezado}: {valor}")

            if partes:
                texto_fila = " | ".join(partes)
                texto_acumulado.append(texto_fila)

                # Agrupar de a 10 filas por chunk para no crear chunks demasiado pequeños
                if len(texto_acumulado) >= 10:
                    fragmento = f"[Hoja: {nombre_hoja}]\n" + "\n".join(texto_acumulado)
                    chunks.append(Chunk(
                        contenido=fragmento,
                        fuente_nombre=nombre_archivo,
                        fuente_tipo="excel",
                        orden_chunk=len(chunks),
                    ))
                    texto_acumulado = []

        # Procesar filas restantes
        if texto_acumulado:
            fragmento = f"[Hoja: {nombre_hoja}]\n" + "\n".join(texto_acumulado)
            chunks.append(Chunk(
                contenido=fragmento,
                fuente_nombre=nombre_archivo,
                fuente_tipo="excel",
                orden_chunk=len(chunks),
            ))

    logger.info(f"Excel '{nombre_archivo}': {len(chunks)} chunks generados")
    return chunks


async def extraer_texto_url(url: str, profundidad_max: int = 2, max_paginas: int = 15) -> list[Chunk]:
    """
    Extrae y limpia el texto de una página web y sus enlaces internos (recursivo).
    """
    import urllib.parse
    logger.info(f"Iniciando scraping recursivo: {url}")

    visitadas = set()
    cola = [(url, 1)]
    chunks = []
    dominio_base = urllib.parse.urlparse(url).netloc

    async with httpx.AsyncClient(
        timeout=httpx.Timeout(15.0),
        follow_redirects=True,
        headers={"User-Agent": "FluxAgent-Bot/2.0 (Knowledge Indexer)"},
    ) as cliente:
        while cola and len(visitadas) < max_paginas:
            actual_url, profundidad = cola.pop(0)
            
            actual_url = actual_url.split('#')[0]
            if actual_url in visitadas:
                continue
                
            visitadas.add(actual_url)
            logger.info(f"Scrapeando URL (prof {profundidad}): {actual_url}")

            try:
                respuesta = await cliente.get(actual_url)
                respuesta.raise_for_status()
                html = respuesta.text
                
                texto = _html_a_texto(html)
                texto = _limpiar_texto(texto)
                
                if texto.strip():
                    fragmentos = _dividir_en_chunks(texto)
                    for i, frag in enumerate(fragmentos):
                        chunks.append(Chunk(
                            contenido=frag,
                            fuente_nombre=url, # Usamos la URL base como name de fuente para agrupar
                            fuente_tipo="url",
                            fuente_url=actual_url,
                            orden_chunk=len(chunks) + i,
                        ))

                if profundidad < profundidad_max:
                    enlaces = set(re.findall(r'href=[\'"]?(https?://[^\'" >]+|/[^\'" >]+)', html))
                    for enlace in enlaces:
                        if enlace.startswith('/'):
                            enlace_abs = urllib.parse.urljoin(actual_url, enlace)
                        else:
                            enlace_abs = enlace
                            
                        if urllib.parse.urlparse(enlace_abs).netloc == dominio_base:
                            if enlace_abs not in visitadas:
                                cola.append((enlace_abs, profundidad + 1))
                                
            except Exception as e:
                logger.warning(f"Error scrapeando {actual_url}: {e}")

    logger.info(f"URL '{url}': {len(chunks)} chunks generados (de {len(visitadas)} páginas)")
    return chunks


def extraer_texto_txt(contenido: bytes, nombre_archivo: str) -> list[Chunk]:
    """Extrae texto de un archivo plano .txt"""
    texto = _limpiar_texto(contenido.decode("utf-8", errors="ignore"))
    if not texto.strip():
        return []
    
    fragmentos = _dividir_en_chunks(texto)
    chunks = [
        Chunk(
            contenido=frag,
            fuente_nombre=nombre_archivo,
            fuente_tipo="texto",
            orden_chunk=i,
        )
        for i, frag in enumerate(fragmentos)
    ]
    logger.info(f"TXT '{nombre_archivo}': {len(chunks)} chunks generados")
    return chunks


def extraer_texto_csv(contenido: bytes, nombre_archivo: str) -> list["Chunk"]:
    """
    Extrae texto de un CSV convirtiendo cada fila en texto narrativo.
    Estrategia idéntica al Excel: cada grupo de 10 filas = un chunk semántico.
    Migrado de v1 (rag_engine_optimized.py) y adaptado al esquema v2.
    """
    import csv

    chunks: list[Chunk] = []
    try:
        texto = contenido.decode("utf-8", errors="ignore")
        lector = csv.DictReader(StringIO(texto))
        texto_acumulado: list[str] = []

        for fila in lector:
            partes = [f"{k}: {v}" for k, v in fila.items() if v and str(v).strip()]
            if partes:
                texto_acumulado.append(" | ".join(partes))

            if len(texto_acumulado) >= 10:
                chunks.append(Chunk(
                    contenido="\n".join(texto_acumulado),
                    fuente_nombre=nombre_archivo,
                    fuente_tipo="csv",
                    orden_chunk=len(chunks),
                ))
                texto_acumulado = []

        if texto_acumulado:
            chunks.append(Chunk(
                contenido="\n".join(texto_acumulado),
                fuente_nombre=nombre_archivo,
                fuente_tipo="csv",
                orden_chunk=len(chunks),
            ))
    except Exception as exc:
        logger.error(f"Error procesando CSV '{nombre_archivo}': {exc}")

    logger.info(f"CSV '{nombre_archivo}': {len(chunks)} chunks generados")
    return chunks


# =============================================================================
# VECTORIZACIÓN (Ollama nomic-embed-text)
# =============================================================================

async def generar_embedding(texto: str, usar_cache: bool = True) -> list[float]:
    """
    Genera un vector de embedding usando Ollama (nomic-embed-text, 768 dims).
    Incluye caché Redis con TTL de 24h para evitar re-vectorizar el mismo texto.
    """
    # ── Cache Redis ─────────────────────────────────────────────────────────
    cache_key = f"emb:{hashlib.md5(texto[:500].encode()).hexdigest()}"
    if usar_cache:
        try:
            import redis.asyncio as aioredis
            r = aioredis.from_url(config.redis_url, decode_responses=True)
            cached = await r.get(cache_key)
            await r.aclose()
            if cached:
                logger.debug("Cache HIT embedding")
                return json.loads(cached)
        except Exception:
            pass  # Redis opcional — degradar a sin cache

    # ── Llamada a Ollama ─────────────────────────────────────────────────────
    async with httpx.AsyncClient(
        base_url=config.ollama_base_url,
        timeout=httpx.Timeout(60.0),
    ) as cliente:
        respuesta = await cliente.post(
            "/api/embeddings",
            json={
                "model": config.ollama_modelo_embedding,
                "prompt": texto[:8000],
            },
        )
        respuesta.raise_for_status()
        datos = respuesta.json()

        embedding = datos.get("embedding")
        if not embedding:
            raise ValueError("Ollama no retornó embedding válido")

    # ── Guardar en cache ─────────────────────────────────────────────────────
    if usar_cache:
        try:
            import redis.asyncio as aioredis
            r = aioredis.from_url(config.redis_url, decode_responses=True)
            await r.setex(cache_key, 86400, json.dumps(embedding))  # TTL 24h
            await r.aclose()
        except Exception:
            pass

    return embedding


# =============================================================================
# PERSISTENCIA EN POSTGRESQL
# =============================================================================

async def guardar_chunks(
    sesion:    AsyncSession,
    chunks:    list[Chunk],
    tenant_id: UUID,
    agent_id:  Optional[UUID] = None,
) -> int:
    """
    Vectoriza y guarda una lista de chunks en la tabla knowledge_chunks.

    Proceso para cada chunk:
      1. Genera el embedding con Ollama
      2. Inserta el registro con el vector en PostgreSQL (pgvector)

    Args:
        sesion    : Sesión activa de SQLAlchemy (con RLS configurado)
        chunks    : Lista de chunks a guardar
        tenant_id : UUID del tenant propietario
        agent_id  : UUID del agente al que pertenece el conocimiento

    Returns:
        Cantidad de chunks guardados exitosamente
    """
    guardados = 0

    for i, chunk in enumerate(chunks):
        try:
            embedding = await generar_embedding(chunk.contenido)

            # Convertir el vector a formato que pgvector entiende: '[0.1, 0.2, ...]'
            vector_str = "[" + ",".join(f"{v:.6f}" for v in embedding) + "]"

            await sesion.execute(
                text("""
                    INSERT INTO knowledge_chunks
                        (tenant_id, agent_id, contenido, fuente_nombre, fuente_tipo,
                         fuente_url, pagina_numero, orden_chunk, tokens_count, embedding)
                    VALUES
                        (:tenant_id, :agent_id, :contenido, :fuente_nombre, :fuente_tipo,
                         :fuente_url, :pagina_numero, :orden_chunk, :tokens_count,
                         CAST(:embedding AS vector))
                """),
                {
                    "tenant_id":     str(tenant_id),
                    "agent_id":      str(agent_id) if agent_id else None,
                    "contenido":     chunk.contenido,
                    "fuente_nombre": chunk.fuente_nombre,
                    "fuente_tipo":   chunk.fuente_tipo,
                    "fuente_url":    chunk.fuente_url,
                    "pagina_numero": chunk.pagina_numero,
                    "orden_chunk":   chunk.orden_chunk,
                    "tokens_count":  len(chunk.contenido.split()),
                    "embedding":     vector_str,
                },
            )
            guardados += 1
            logger.debug(f"Chunk {i+1}/{len(chunks)} guardado: {chunk.fuente_nombre[:30]}...")

        except Exception as exc:
            logger.error(f"Error guardando chunk {i}: {exc}")
            # Continuar con el siguiente chunk en lugar de abortar todo

    logger.info(f"Ingesta completada: {guardados}/{len(chunks)} chunks guardados")
    return guardados


# =============================================================================
# BÚSQUEDA SEMÁNTICA
# =============================================================================

async def buscar_chunks_relevantes(
    sesion:    AsyncSession,
    consulta:  str,
    tenant_id: UUID,
    agent_id:  Optional[UUID] = None,
    top_k:     int = 5,
) -> list[dict]:
    """
    Busca los chunks más relevantes para una consulta usando similitud coseno.

    Flujo:
      1. Vectoriza la consulta del usuario con el mismo model de embedding
      2. Calcula la similitud coseno contra todos los chunks del tenant
      3. Retorna los top_k más similares

    Args:
        sesion    : Sesión activa de SQLAlchemy (RLS aplicado)
        consulta  : Pregunta o mensaje del usuario
        tenant_id : UUID del tenant (RLS garantiza aislamiento)
        agent_id  : Si se especifica, filtra solo los chunks de ese agente
        top_k     : Número máximo de chunks a retornar

    Returns:
        Lista de dicts con 'contenido' y 'similitud'
    """
    embedding_consulta = await generar_embedding(consulta)
    vector_str = "[" + ",".join(f"{v:.6f}" for v in embedding_consulta) + "]"

    # Filtro opcional por agent_id
    filtro_agente = "AND agent_id = :agent_id" if agent_id else ""

    resultado = await sesion.execute(
        text(f"""
            SELECT
                contenido,
                fuente_nombre,
                1 - (embedding <=> CAST(:embedding AS vector)) AS similitud
            FROM knowledge_chunks
            WHERE tenant_id = :tenant_id
              {filtro_agente}
              AND embedding IS NOT NULL
            ORDER BY embedding <=> CAST(:embedding AS vector)
            LIMIT :top_k
        """),
        {
            "embedding":  vector_str,
            "tenant_id":  str(tenant_id),
            "agent_id":   str(agent_id) if agent_id else None,
            "top_k":      top_k,
        },
    )

    filas = resultado.fetchall()
    chunks_relevantes = [
        {
            "contenido":     fila.contenido,
            "fuente_nombre": fila.fuente_nombre,
            "similitud":     round(float(fila.similitud), 4),
        }
        for fila in filas
        if fila.similitud > 0.3   # Umbral mínimo de relevancia
    ]

    logger.debug(f"RAG: {len(chunks_relevantes)} chunks relevantes encontrados para '{consulta[:50]}...'")
    return chunks_relevantes


# =============================================================================
# CLASE ORQUESTADORA (Fachada pública del módulo)
# =============================================================================

class ServicioIngesta:
    """
    Fachada que unifica el pipeline completo de ingesta RAG.
    Es el punto de entrada principal para los endpoints de FastAPI.
    """

    async def procesar_pdf(
        self,
        sesion:    AsyncSession,
        contenido: bytes,
        name:    str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
    ) -> int:
        """Procesa un PDF y persiste sus chunks vectorizados."""
        chunks = extraer_texto_pdf(contenido, name)
        return await guardar_chunks(sesion, chunks, tenant_id, agent_id)

    async def procesar_excel(
        self,
        sesion:    AsyncSession,
        contenido: bytes,
        name:    str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
    ) -> int:
        """Procesa un Excel y persiste sus chunks vectorizados."""
        import openpyxl
        from io import BytesIO
        from sqlalchemy import text
        try:
            wb = openpyxl.load_workbook(BytesIO(contenido), read_only=True, data_only=True)
            for nombre_hoja in wb.sheetnames:
                hoja = wb[nombre_hoja]
                filas = list(hoja.iter_rows(values_only=True))
                if not filas: continue
                encabezados = [str(c).strip().lower() if c else f"columna_{i}" for i, c in enumerate(filas[0])]
                
                # Buscar indices
                idx_nombre = next((i for i, h in enumerate(encabezados) if "name" in h or "producto" in h or "name" in h or "articulo" in h), -1)
                idx_precio = next((i for i, h in enumerate(encabezados) if "price" in h or "price" in h or "valor" in h), -1)
                idx_stock = next((i for i, h in enumerate(encabezados) if "stock" in h or "cantidad" in h or "qty" in h), -1)
                
                if idx_nombre >= 0:
                    for fila in filas[1:]:
                        nom = str(fila[idx_nombre]) if fila[idx_nombre] else "Desconocido"
                        pre = fila[idx_precio] if idx_precio >= 0 and fila[idx_precio] else 0.0
                        stk = fila[idx_stock] if idx_stock >= 0 and fila[idx_stock] else 0
                        
                        try: pre_val = float(pre)
                        except: pre_val = 0.0
                        try: stk_val = int(stk)
                        except: stk_val = 0
                        
                        await sesion.execute(text("""
                            INSERT INTO productos (tenant_id, name, price, stock)
                            VALUES (:tid, :nom, :pre, :stk)
                        """), {"tid": str(tenant_id), "nom": nom[:250], "pre": pre_val, "stk": stk_val})
        except Exception as e:
            logger.warning(f"No se pudieron extraer productos estructurados del Excel: {e}")

        chunks = extraer_texto_excel(contenido, name)
        return await guardar_chunks(sesion, chunks, tenant_id, agent_id)

    async def procesar_txt(
        self,
        sesion:    AsyncSession,
        contenido: bytes,
        name:    str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
    ) -> int:
        """Procesa un TXT y persiste sus chunks vectorizados."""
        chunks = extraer_texto_txt(contenido, name)
        return await guardar_chunks(sesion, chunks, tenant_id, agent_id)

    async def procesar_csv(
        self,
        sesion:    AsyncSession,
        contenido: bytes,
        name:    str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
    ) -> int:
        """Procesa un CSV y persiste sus chunks vectorizados."""
        import csv
        from io import StringIO
        from sqlalchemy import text
        try:
            texto = contenido.decode("utf-8", errors="ignore")
            lector = csv.DictReader(StringIO(texto))
            for fila in lector:
                # buscar heuristicamente
                nombres = [v for k, v in fila.items() if k and ("name" in k.lower() or "producto" in k.lower() or "name" in k.lower() or "articulo" in k.lower())]
                precios = [v for k, v in fila.items() if k and ("price" in k.lower() or "price" in k.lower() or "valor" in k.lower())]
                stocks = [v for k, v in fila.items() if k and ("stock" in k.lower() or "cantidad" in k.lower() or "qty" in k.lower())]
                
                nom = nombres[0] if nombres else "Desconocido"
                pre = precios[0] if precios else 0.0
                stk = stocks[0] if stocks else 0
                
                try: pre_val = float(pre)
                except: pre_val = 0.0
                try: stk_val = int(stk)
                except: stk_val = 0
                
                await sesion.execute(text("""
                    INSERT INTO productos (tenant_id, name, price, stock)
                    VALUES (:tid, :nom, :pre, :stk)
                """), {"tid": str(tenant_id), "nom": str(nom)[:250], "pre": pre_val, "stk": stk_val})
        except Exception as e:
            logger.warning(f"No se pudieron extraer productos estructurados del CSV: {e}")

        chunks = extraer_texto_csv(contenido, name)
        return await guardar_chunks(sesion, chunks, tenant_id, agent_id)

    async def procesar_url(
        self,
        sesion:    AsyncSession,
        url:       str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
    ) -> int:
        """Scrapea una URL y persiste sus chunks vectorizados."""
        chunks = await extraer_texto_url(url)
        return await guardar_chunks(sesion, chunks, tenant_id, agent_id)

    async def buscar(
        self,
        sesion:    AsyncSession,
        consulta:  str,
        tenant_id: UUID,
        agent_id:  Optional[UUID] = None,
        top_k:     int = 5,
    ) -> list[dict]:
        """Búsqueda semántica en la base de conocimiento del tenant."""
        return await buscar_chunks_relevantes(sesion, consulta, tenant_id, agent_id, top_k)


# =============================================================================
# HELPERS PRIVADOS
# =============================================================================

def _limpiar_texto(texto: str) -> str:
    """Elimina caracteres de control, espacios múltiples y líneas vacías."""
    texto = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', texto)
    texto = re.sub(r'\n{3,}', '\n\n', texto)
    texto = re.sub(r' {2,}', ' ', texto)
    return texto.strip()


def _dividir_en_chunks(texto: str, tamano: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """
    Divide texto largo en fragmentos con solapamiento para preservar contexto.
    Intenta respetar límites de párrafo cuando es posible.
    """
    if len(texto) <= tamano:
        return [texto]

    parrafos = texto.split('\n\n')
    chunks = []
    acumulado = ""

    for parrafo in parrafos:
        if len(acumulado) + len(parrafo) <= tamano:
            acumulado += ("\n\n" if acumulado else "") + parrafo
        else:
            if acumulado:
                chunks.append(acumulado)
                # Solapamiento: incluir el final del chunk anterior
                acumulado = acumulado[-overlap:] + "\n\n" + parrafo if overlap else parrafo
            else:
                # Párrafo más largo que el tamaño de chunk — dividir por líneas
                chunks.extend(wrap(parrafo, tamano))
                acumulado = ""

    if acumulado:
        chunks.append(acumulado)

    return [c for c in chunks if c.strip()]


def _html_a_texto(html: str) -> str:
    """
    Convierte HTML a texto plano eliminando etiquetas, scripts y estilos.
    Sin dependencias externas — usa sólo regex.
    """
    # Eliminar scripts y estilos completos
    html = re.sub(r'<script[^>]*>.*?</script>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<style[^>]*>.*?</style>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    # Convertir saltos de línea HTML en reales
    html = re.sub(r'<br\s*/?>', '\n', html, flags=re.IGNORECASE)
    html = re.sub(r'</p>|</div>|</h[1-6]>|</li>', '\n', html, flags=re.IGNORECASE)
    # Eliminar resto de etiquetas
    html = re.sub(r'<[^>]+>', ' ', html)
    # Decodificar entidades HTML básicas
    html = html.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')
    return html
